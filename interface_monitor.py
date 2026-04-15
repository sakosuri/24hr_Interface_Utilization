"""
Interface Monitoring Automation Script
=======================================
This script:
1. Reads device and interface details from a YAML configuration file.
2. Queries Catalyst Center (DNA Center) APIs to retrieve interface status
   (admin status and operational status).
3. Queries the Trend Analytics API to retrieve Min/Max Tx/Rx rates
   over the last 24 hours with a 10-minute aggregation interval.
4. Exports all results to an Excel file.
"""

import yaml
import requests
import urllib3
import time
import sys
import os
from datetime import datetime, timedelta, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Suppress insecure HTTPS warnings (common for lab/internal Catalyst Center instances)
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ──────────────────────────────────────────────
# Constants
# ──────────────────────────────────────────────
TOKEN_API_PATH = "/dna/system/api/v1/auth/token"
INTERFACES_API_PATH = "/dna/data/api/v1/interfaces"
TREND_ANALYTICS_API_PATH = "/dna/data/api/v1/interfaces/{id}/trendAnalytics"
TREND_INTERVAL_MINUTES = 10
LOOKBACK_HOURS = 24
REQUEST_TIMEOUT = 60  # seconds


# ──────────────────────────────────────────────
# Configuration Loader
# ──────────────────────────────────────────────
def load_config(config_path: str) -> dict:
    """Load and return the YAML configuration file."""
    if not os.path.exists(config_path):
        print(f"[ERROR] Configuration file not found: {config_path}")
        sys.exit(1)

    with open(config_path, "r") as f:
        config = yaml.safe_load(f)

    # Basic validation
    if not config.get("dna_centers"):
        print("[ERROR] 'dna_centers' section is missing in config.yaml")
        sys.exit(1)
    if not config.get("targets"):
        print("[ERROR] 'targets' section is missing in config.yaml")
        sys.exit(1)

    return config


# ──────────────────────────────────────────────
# Authentication
# ──────────────────────────────────────────────
def get_auth_token(base_url: str, username: str, password: str) -> str:
    """
    Authenticate with Catalyst Center and return a reusable token.

    Parameters:
        base_url : str  – e.g. "https://10.1.1.1"
        username : str
        password : str

    Returns:
        str – authentication token
    """
    url = f"{base_url}{TOKEN_API_PATH}"
    print(f"[INFO] Authenticating with Catalyst Center at {base_url} ...")

    try:
        response = requests.post(
            url,
            auth=(username, password),
            headers={"Content-Type": "application/json"},
            verify=False,
            timeout=REQUEST_TIMEOUT,
        )
        response.raise_for_status()
        token = response.json().get("Token")
        if not token:
            print("[ERROR] Token not found in authentication response.")
            sys.exit(1)
        print("[INFO] Authentication successful.")
        return token

    except requests.exceptions.RequestException as e:
        print(f"[ERROR] Authentication failed: {e}")
        sys.exit(1)


# ──────────────────────────────────────────────
# API 1 – Get Interface Details
# ──────────────────────────────────────────────
def get_interface_details(
    base_url: str, token: str, device_ip: str, interface_name: str
) -> dict:
    """
    Retrieve interface details (adminStatus, operStatus, id, etc.)
    for a specific device IP and interface name.

    Parameters:
        base_url       : str – e.g. "https://10.1.1.1"
        token          : str – auth token
        device_ip      : str – management IP of the device
        interface_name : str – e.g. "GigabitEthernet1/0/1"

    Returns:
        dict – interface record from the API response, or None
    """
    url = f"{base_url}{INTERFACES_API_PATH}"
    headers = {
        "X-Auth-Token": token,
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    params = {
        "networkDeviceIpAddress": device_ip,
        "interfaceName": interface_name,
    }

    print(
        f"[INFO] Fetching interface details for device={device_ip}, "
        f"interface={interface_name} ..."
    )

    try:
        response = requests.get(
            url,
            headers=headers,
            params=params,
            verify=False,
            timeout=REQUEST_TIMEOUT,
        )
        response.raise_for_status()
        data = response.json()

        interfaces = data.get("response", [])
        if not interfaces:
            print(
                f"[WARN] No interface data returned for device={device_ip}, "
                f"interface={interface_name}"
            )
            return None

        # Return the first matching record
        interface_record = interfaces[0]
        print(
            f"  -> id={interface_record.get('id')}, "
            f"adminStatus={interface_record.get('adminStatus')}, "
            f"operStatus={interface_record.get('operStatus')}, "
            f"speed={interface_record.get('speed')}, "
            f"duplex_oper={interface_record.get('duplexOper')}, "
            f"duplex_config={interface_record.get('duplexConfig')}"
        )
        return interface_record

    except requests.exceptions.RequestException as e:
        print(
            f"[ERROR] Failed to fetch interface details for device={device_ip}, "
            f"interface={interface_name}: {e}"
        )
        return None


# ──────────────────────────────────────────────
# API 2 – Get Trend Analytics (Min/Max Tx/Rx)
# ──────────────────────────────────────────────
def get_trend_analytics(base_url: str, token: str, interface_id: str) -> dict:
    """
    Retrieve trend analytics (Min/Max of TXRATE and RXRATE) for the
    last 24 hours with a 10-minute aggregation interval.

    Parameters:
        base_url     : str – e.g. "https://10.1.1.1"
        token        : str – auth token
        interface_id : str – interface instance UUID

    Returns:
        dict – {"min_tx": value, "min_rx": value, "max_tx": value, "max_rx": value}
               or None on failure
    """
    url = f"{base_url}{TREND_ANALYTICS_API_PATH.format(id=interface_id)}"
    headers = {
        "X-Auth-Token": token,
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    params = {
        "id": interface_id
    }

    # Calculate epoch milliseconds for the last 24 hours
    now = datetime.now(timezone.utc)
    end_time = int(now.timestamp() * 1000)
    start_time = int((now - timedelta(hours=LOOKBACK_HOURS)).timestamp() * 1000)

    request_body = {
        "startTime": start_time,
        "endTime": end_time,
        "trendIntervalInMinutes": TREND_INTERVAL_MINUTES,
        "aggregateAttributes": [
            {"name": "txRate", "function": "min"},
            {"name": "txRate", "function": "max"},
            {"name": "rxRate", "function": "min"},
            {"name": "rxRate", "function": "max"},
            {"name": "rxError", "function": "max"},
            {"name": "rxError", "function": "min"},
            {"name": "txError", "function": "max"},
            {"name": "txError", "function": "min"},
            {"name": "txDiscards", "function": "max"},
            {"name": "txDiscards", "function": "min"},
            {"name": "rxDiscards", "function": "max"},
            {"name": "rxDiscards", "function": "min"},
            {"name": "rxUtilization", "function": "max"},
            {"name": "rxUtilization", "function": "min"},
            {"name": "txUtilization", "function": "max"},
            {"name": "txUtilization", "function": "min"}

        ]
    }

    print(
        f"[INFO] Fetching trend analytics for interface_id={interface_id} "
        f"(last {LOOKBACK_HOURS}h, interval={TREND_INTERVAL_MINUTES}min) ..."
    )

    try:
        response = requests.post(
            url,
            headers=headers,
            json=request_body,
            params=params,
            verify=False,
            timeout=REQUEST_TIMEOUT
        )
        response.raise_for_status()
        data = response.json()

        trend_entries = data.get("response", [])
        if not trend_entries:
            print(
                f"[WARN] No trend analytics data returned for "
                f"interface_id={interface_id}"
            )
            return None

        # ── Aggregate across all time-interval buckets ──
        # Collect all MIN TXRATE values, all MAX TXRATE values, etc.
        all_min_tx = []
        all_max_tx = []
        all_min_rx = []
        all_max_rx = []
        all_min_tx_error = []
        all_max_tx_error = []
        all_min_rx_error = []
        all_max_rx_error = []
        all_min_tx_discards = []
        all_max_tx_discards = []
        all_min_rx_discards = []
        all_max_rx_discards = []
        all_min_rx_utilization = []
        all_max_rx_utilization = []
        all_min_tx_utilization = []
        all_max_tx_utilization = []

        for entry in trend_entries:
            aggregate_attrs = entry.get("aggregateAttributes", [])
            for attr in aggregate_attrs:
                name = attr.get("name", "").upper()
                function = attr.get("function", "").upper()
                value = attr.get("value")

                if value is None:
                    continue

                if name == "TXRATE" and function == "MIN":
                    all_min_tx.append(value)
                elif name == "TXRATE" and function == "MAX":
                    all_max_tx.append(value)
                elif name == "RXRATE" and function == "MIN":
                    all_min_rx.append(value)
                elif name == "RXRATE" and function == "MAX":
                    all_max_rx.append(value)
                elif name == "RXERROR" and function == "MIN":
                    all_min_rx_error.append(value)
                elif name == "RXERROR" and function == "MAX":
                    all_max_rx_error.append(value)
                elif name == "TXERROR" and function == "MIN":
                    all_min_tx_error.append(value)
                elif name == "TXERROR" and function == "MAX":
                    all_max_tx_error.append(value)
                elif name == "TXDISCARDS" and function == "MIN":
                    all_min_tx_discards.append(value)
                elif name == "TXDISCARDS" and function == "MAX":
                    all_max_tx_discards.append(value)
                elif name == "RXDISCARDS" and function == "MIN":
                    all_min_rx_discards.append(value)
                elif name == "RXDISCARDS" and function == "MAX":
                    all_max_rx_discards.append(value)
                elif name == "RXUTILIZATION" and function == "MIN":
                    all_min_rx_utilization.append(value)
                elif name == "RXUTILIZATION" and function == "MAX":
                    all_max_rx_utilization.append(value)
                elif name == "TXUTILIZATION" and function == "MIN":
                    all_min_tx_utilization.append(value)
                elif name == "TXUTILIZATION" and function == "MAX":
                    all_max_tx_utilization.append(value)

        result = {
            "min_tx": min(all_min_tx) if all_min_tx else "N/A",
            "min_rx": min(all_min_rx) if all_min_rx else "N/A",
            "max_tx": max(all_max_tx) if all_max_tx else "N/A",
            "max_rx": max(all_max_rx) if all_max_rx else "N/A",
            "min_tx_error": min(all_min_tx_error) if all_min_tx_error else "N/A",
            "max_tx_error": max(all_max_tx_error) if all_max_tx_error else "N/A",
            "min_rx_error": min(all_min_rx_error) if all_min_rx_error else "N/A",
            "max_rx_error": max(all_max_rx_error) if all_max_rx_error else "N/A",
            "min_tx_discards": min(all_min_tx_discards) if all_min_tx_discards else "N/A",
            "max_tx_discards": max(all_max_tx_discards) if all_max_tx_discards else "N/A",
            "min_rx_discards": min(all_min_rx_discards) if all_min_rx_discards else "N/A",
            "max_rx_discards": max(all_max_rx_discards) if all_max_rx_discards else "N/A",
            "min_rx_utilization": min(all_min_rx_utilization) if all_min_rx_utilization else "N/A",
            "max_rx_utilization": max(all_max_rx_utilization) if all_max_rx_utilization else "N/A",
            "min_tx_utilization": min(all_min_tx_utilization) if all_min_tx_utilization else "N/A",
            "max_tx_utilization": max(all_max_tx_utilization) if all_max_tx_utilization else "N/A"
        }

        print(
            f"  -> min_tx={result['min_tx']}, min_rx={result['min_rx']}, "
            f"max_tx={result['max_tx']}, max_rx={result['max_rx']}, "
            f"min_tx_error={result['min_tx_error']}, max_tx_error={result['max_tx_error']}, "
            f"min_rx_error={result['min_rx_error']}, max_rx_error={result['max_rx_error']}, "
            f"min_tx_discards={result['min_tx_discards']}, max_tx_discards={result['max_tx_discards']}, "
            f"min_rx_discards={result['min_rx_discards']}, max_rx_discards={result['max_rx_discards']}, "
            f"min_rx_utilization={result['min_rx_utilization']}, max_rx_utilization={result['max_rx_utilization']}, "
            f"min_tx_utilization={result['min_tx_utilization']}, max_tx_utilization={result['max_tx_utilization']}"
        )
        return result

    except requests.exceptions.RequestException as e:
        print(
            f"[ERROR] Failed to fetch trend analytics for "
            f"interface_id={interface_id}: {e}"
        )
        return None


# ──────────────────────────────────────────────
# Excel Export
# ──────────────────────────────────────────────
def export_to_excel(results: list, output_file: str) -> None:
    """
    Export the collected results to a formatted Excel file.

    Parameters:
        results     : list of dicts with keys:
                      device_ip, interface_name, admin_status, oper_status,
                      min_tx, min_rx, max_tx, max_rx
        output_file : str – output .xlsx file path
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Interface Report"

    # ── Header styling ──
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ── Define headers ──
    headers = [
        "Device IP Address",
        "Interface Name",
        "Admin Status",
        "Oper Status",
        "Speed (Kbps)",
        "Duplex (Oper)",
        "Duplex (Config)",
        "Min Tx (bps)",
        "Min Rx (bps)",
        "Max Tx (bps)",
        "Max Rx (bps)",
        "Min Tx Errors (%)",
        "Max Tx Errors (%)",
        "Min Rx Errors (%)",
        "Max Rx Errors (%)",
        "Min Tx Discards (%)",
        "Max Tx Discards (%)",
        "Min Rx Discards (%)",
        "Max Rx Discards (%)",
        "Min Rx Utilization (%)",
        "Max Rx Utilization (%)",
        "Min Tx Utilization (%)",
        "Max Tx Utilization (%)"
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # ── Write data rows ──
    data_alignment = Alignment(horizontal="center", vertical="center")
    for row_num, record in enumerate(results, 2):
        row_data = [
            record.get("device_ip", "N/A"),
            record.get("interface_name", "N/A"),
            record.get("admin_status", "N/A"),
            record.get("oper_status", "N/A"),
            record.get("speed", "N/A"),
            record.get("duplex_oper", "N/A"),
            record.get("duplex_config", "N/A"),
            record.get("min_tx", "N/A"),
            record.get("min_rx", "N/A"),
            record.get("max_tx", "N/A"),
            record.get("max_rx", "N/A"),
            record.get("min_tx_error", "N/A"),
            record.get("max_tx_error", "N/A"),
            record.get("min_rx_error", "N/A"),
            record.get("max_rx_error", "N/A"),
            record.get("min_tx_discards", "N/A"),
            record.get("max_tx_discards", "N/A"),
            record.get("min_rx_discards", "N/A"),
            record.get("max_rx_discards", "N/A"),
            record.get("min_rx_utilization", "N/A"),
            record.get("max_rx_utilization", "N/A"),
            record.get("min_tx_utilization", "N/A"),
            record.get("max_tx_utilization", "N/A")
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = data_alignment
            cell.border = thin_border

    # ── Auto-adjust column widths ──
    for col_num, header in enumerate(headers, 1):
        max_length = len(header)
        for row_num in range(2, len(results) + 2):
            cell_value = str(ws.cell(row=row_num, column=col_num).value or "")
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[
            ws.cell(row=1, column=col_num).column_letter
        ].width = max_length + 4

    # ── Save ──
    wb.save(output_file)
    print(f"\n[INFO] Report exported successfully to: {output_file}")


# ──────────────────────────────────────────────
# Main Orchestrator
# ──────────────────────────────────────────────
def main():
    """Main entry point that orchestrates the full workflow."""
    config_path = "config.yaml"
    output_file = (
        f"interface_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    # ── Step 1: Load configuration ──
    print("=" * 70)
    print("  Interface Monitoring Automation Script")
    print("=" * 70)
    config = load_config(config_path)

    # Build a lookup map for DNA Center instances
    dnac_map = {}
    for dnac in config["dna_centers"]:
        dnac_map[dnac["name"]] = {
            "ip": dnac["ip"],
            "username": dnac["username"],
            "password": dnac["password"],
            "base_url": f"https://{dnac['ip']}",
        }

    # ── Step 2: Iterate through targets ──
    all_results = []

    for target in config["targets"]:
        dnac_name = target["dna_center_name"]
        if dnac_name not in dnac_map:
            print(
                f"[ERROR] DNA Center '{dnac_name}' referenced in targets "
                f"but not defined in dna_centers. Skipping."
            )
            continue

        dnac_info = dnac_map[dnac_name]
        base_url = dnac_info["base_url"]

        # Authenticate once per DNA Center
        token = get_auth_token(
            base_url, dnac_info["username"], dnac_info["password"]
        )

        for device in target.get("devices", []):
            device_ip = device["device_ip"]

            for interface_name in device.get("interfaces", []):
                print("-" * 50)

                # ── API 1: Get interface admin/oper status and ID ──
                interface_data = get_interface_details(
                    base_url, token, device_ip, interface_name
                )

                if interface_data is None:
                    # Record with N/A values if interface not found
                    all_results.append(
                        {
                            "device_ip": device_ip,
                            "interface_name": interface_name,
                            "admin_status": "N/A",
                            "oper_status": "N/A",
                            "speed": "N/A",
                            "duplex_oper": "N/A",
                            "duplex_config": "N/A",
                            "min_tx": "N/A",
                            "min_rx": "N/A",
                            "max_tx": "N/A",
                            "max_rx": "N/A",
                            "min_tx_error": "N/A",
                            "max_tx_error": "N/A",
                            "min_rx_error": "N/A",
                            "max_rx_error": "N/A",
                            "min_tx_discards": "N/A",
                            "max_tx_discards": "N/A",
                            "min_rx_discards": "N/A",
                            "max_rx_discards": "N/A",
                            "min_rx_utilization": "N/A",
                            "max_rx_utilization": "N/A",
                            "min_tx_utilization": "N/A",
                            "max_tx_utilization": "N/A",
                        }
                    )
                    continue

                admin_status = interface_data.get("adminStatus", "N/A")
                oper_status = interface_data.get("operStatus", "N/A")
                interface_id = interface_data.get("id")
                speed = interface_data.get("speed", "N/A")
                duplex_oper = interface_data.get("duplexOper", "N/A")
                duplex_config = interface_data.get("duplexConfig", "N/A")

                if not interface_id:
                    print(
                        f"[WARN] Interface ID not found for {interface_name} "
                        f"on {device_ip}. Skipping trend analytics."
                    )
                    all_results.append(
                        {
                            "device_ip": device_ip,
                            "interface_name": interface_name,
                            "admin_status": admin_status,
                            "oper_status": oper_status,
                            "speed": speed,
                            "duplex_oper": duplex_oper,
                            "duplex_config": duplex_config,
                            "min_tx": "N/A",
                            "min_rx": "N/A",
                            "max_tx": "N/A",
                            "max_rx": "N/A",
                            "min_tx_error": "N/A",
                            "max_tx_error": "N/A",
                            "min_rx_error": "N/A",
                            "max_rx_error": "N/A",
                            "min_tx_discards": "N/A",
                            "max_tx_discards": "N/A",
                            "min_rx_discards": "N/A",
                            "max_rx_discards": "N/A",
                            "min_rx_utilization": "N/A",
                            "max_rx_utilization": "N/A",
                            "min_tx_utilization": "N/A",
                            "max_tx_utilization": "N/A",
                        }
                    )
                    continue

                # ── API 2: Get trend analytics (Min/Max Tx/Rx) ──
                trend_data = get_trend_analytics(base_url, token, interface_id)

                if trend_data is None:
                    trend_data = {
                        "min_tx": "N/A",
                        "min_rx": "N/A",
                        "max_tx": "N/A",
                        "max_rx": "N/A",
                        "min_tx_error": "N/A",
                        "max_tx_error": "N/A",
                        "min_rx_error": "N/A",
                        "max_rx_error": "N/A",
                        "min_tx_discards": "N/A",
                        "max_tx_discards": "N/A",
                        "min_rx_discards": "N/A",
                        "max_rx_discards": "N/A",
                        "min_rx_utilization": "N/A",
                        "max_rx_utilization": "N/A",
                        "min_tx_utilization": "N/A",
                        "max_tx_utilization": "N/A",
                    }

                all_results.append(
                    {
                        "device_ip": device_ip,
                        "interface_name": interface_name,
                        "admin_status": admin_status,
                        "oper_status": oper_status,
                        "speed": speed,
                        "duplex_oper": duplex_oper,
                        "duplex_config": duplex_config,
                        "min_tx": trend_data["min_tx"],
                        "min_rx": trend_data["min_rx"],
                        "max_tx": trend_data["max_tx"],
                        "max_rx": trend_data["max_rx"],
                        "min_tx_error": trend_data["min_tx_error"],
                        "max_tx_error": trend_data["max_tx_error"],
                        "min_rx_error": trend_data["min_rx_error"],
                        "max_rx_error": trend_data["max_rx_error"],
                        "min_tx_discards": trend_data["min_tx_discards"],
                        "max_tx_discards": trend_data["max_tx_discards"],
                        "min_rx_discards": trend_data["min_rx_discards"],
                        "max_rx_discards": trend_data["max_rx_discards"],
                        "min_rx_utilization": trend_data["min_rx_utilization"],
                        "max_rx_utilization": trend_data["max_rx_utilization"],
                        "min_tx_utilization": trend_data["min_tx_utilization"],
                        "max_tx_utilization": trend_data["max_tx_utilization"],
                    }
                )
                # Small delay to avoid API rate limiting
                time.sleep(1)

    # ── Step 3: Export results to Excel ──
    print("=" * 70)
    if all_results:
        export_to_excel(all_results, output_file)
    else:
        print("[WARN] No results collected. Excel file will not be generated.")

    # ── Summary ──
    print(f"\n[INFO] Total interfaces processed: {len(all_results)}")
    print("=" * 70)


if __name__ == "__main__":
    main()
